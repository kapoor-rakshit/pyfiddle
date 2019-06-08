# REFERENCE : https://openpyxl.readthedocs.io/en/stable/tutorial.html

from openpyxl import Workbook, load_workbook

# A workbook is always created with at least one worksheet.
#----------------------------------------------------------
wb = Workbook()

# get the current sheet
#----------------------
ws1 = wb.active



# create new worksheets
#-----------------------
# 1
# title="sheet" and insert at the end (default)
ws3 = wb.create_sheet("sheet")

# 2
# title="sheetNumber2" and insert at desired position (0,1,2...)
ws2 = wb.create_sheet("sheetNumber2", 1)



# Sheets are given a name automatically when they are created.
# They are numbered in sequence (Sheet, Sheet1, Sheet2, …)
# We can change name of sheets
#--------------------------------
ws1.title = "Summary"
ws3.title = "Report"



# The background color of the tab holding this title is white by default
# We can change this providing an RRGGBB color code
#---------------------------------------------------
ws1.sheet_properties.tabColor = "1072BA"
ws2.sheet_properties.tabColor = "00FF00"
ws3.sheet_properties.tabColor = "CCDFAA"



# given a worksheet name, we can get it as a key of the workbook
#----------------------------------------------------------------
chk_ws = wb["Summary"]
print(chk_ws.sheet_properties.tabColor)



# create copy of worksheets within a single workbook with title = SourceTitle Copy
#----------------------------------------------------------------------------------
source = ws3
copied_ws = wb.copy_worksheet(source)



# names of all worksheets of the workbook
#-----------------------------------------
# 1
print(wb.sheetnames)

# 2
for sheet in wb:
	print(sheet.title)



# Data -- CRUD
#---------------
# 1
# Cells can be accessed directly as keys of the worksheet
val = ws1['A4']
ws1['A8'] = 14109025

# 2
# access to cells using row and column notation
cell = ws1.cell(row=4, column=2, value=10)
cell.value = 50

# example
for x in range(1,101):
	for y in range(1,101):
		ws2.cell(row=x, column=y, value="_/\\_ Hari Om ji _/\\_")

# 3
# Ranges of cells can be accessed using slicing
# NOTE : When a worksheet is created in memory, it contains no cells. They are created when first accessed.
#        That is why the results are terminated only to respective row/column accessed till now.
cell_range = ws1['A1':'C2']
print(cell_range)                    # ((<Cell 'Summary'.A1>, <Cell 'Summary'.B1>, <Cell 'Summary'.C1>), (<Cell 'Summary'.A2>, <Cell 'Summary'.B2>, <Cell 'Summary'.C2>))

# 4
# Ranges of rows or columns can be obtained
colC = ws1['C']
print(colC)                         # (<Cell 'Summary'.C1>, <Cell 'Summary'.C2>, <Cell 'Summary'.C3>, <Cell 'Summary'.C4>,...........................<Cell 'Summary'.C8>)

col_range = ws1['C:D']
print(col_range)                    # ((<Cell 'Summary'.C1>, <Cell 'Summary'.C2>,....,<Cell 'Summary'.C8>), (<Cell 'Summary'.D1>, <Cell 'Summary'.D2>,....,<Cell 'Summary'.D8>))

row10 = ws1[10]
print(row10)                        # (<Cell 'Summary'.A10>, <Cell 'Summary'.B10>, <Cell 'Summary'.C10>, <Cell 'Summary'.D10>)

row_range = ws1[5:10]
print(row_range)                    # ((<Cell 'Summary'.A5>,..,<Cell 'Summary'.D5>),....,(<Cell 'Summary'.A9>,..,<Cell 'Summary'.D9>), (<Cell 'Summary'.A10>,..,<Cell 'Summary'.D10>))

# 5
# iterate row-wise
for row in ws1.iter_rows(min_row=1, min_col=6, max_col=9, max_row=1):
	for cell in row:
		print(cell)                 # <Cell 'Summary'.F1>, <Cell 'Summary'.G1>, <Cell 'Summary'.H1>, <Cell 'Summary'.I1>

# 6
# iterate column-wise
for col in ws1.iter_cols(min_row=1, max_col=3, max_row=2):
	for cell in col:
		print(cell)

# 7
# all the rows or columns of a file in memory
print(ws1.rows)
print(ws1.columns)

# 8
# values of cells in memory
for row in ws1.values:
	for value in row:
		print(value)



# save workbook with desired name
#---------------------------------
# This operation will overwrite existing files without warning.
wb.save("testXLSX.xlsx")



# Loading Workbook
#------------------
# use the openpyxl.load_workbook() to open an existing workbook
wb2 = load_workbook('test.xlsx')
print(wb2.sheetnames)


